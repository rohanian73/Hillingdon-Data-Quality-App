import json
import math
import tempfile
import uuid

import uvicorn
from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
# from pydantic import BaseModel
# from typing import List
import csv
import io
import pandas as pd
from pandas.api.types import is_string_dtype
# from ydata_profiling import ProfileReport
import re
from fastapi.responses import Response
# import spacy
# from spacy.training.example import Example
from rapidfuzz import process, fuzz
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import classification_report
from collections import Counter
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import IsolationForest
import numpy as np
from sentence_transformers import SentenceTransformer, util
from sklearn.metrics.pairwise import cosine_similarity
import faiss
from sklearn.cluster import AgglomerativeClustering



app = FastAPI()

origins = [
    "http://localhost:3000"
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

highlighted_excel_sheet = None
corrected_excel_sheet = None
corrected_csv_sheet = None


@app.get("/")
def root():
    return {"Hello": "World"}


@app.post("/analyse_data")
async def read_file(file: UploadFile = File(...)):
    global highlighted_excel_sheet
    global corrected_excel_sheet
    global corrected_csv_sheet

    highlighted_excel_sheet = None
    corrected_excel_sheet = None
    corrected_csv_sheet = None

    contents = await file.read()
    fileName = file.filename

    excel_data = None

    if fileName.endswith(".csv"):
        file_type = "csv"
        csv_data = io.StringIO(contents.decode("utf-8"))
        df = pd.read_csv(csv_data, dtype=str)

        excel_buffer = io.BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)
        excel_data = excel_buffer

        wb_corrected = load_workbook(excel_data)
        ws_corrected = wb_corrected.active

        # df_corrected = df.copy()
    elif fileName.endswith(".xlsx"):
        file_type = "xlsx"
        excel_data = io.BytesIO(contents)
        df = pd.read_excel(excel_data, dtype=str)

        wb_corrected = load_workbook(excel_data)
        ws_corrected = wb_corrected.active
        # df_corrected = df.copy()
    else:
        return {"error": "File type not supported."}


    # print(df.info())
    # print(df_corrected.info())
    # print("ILOC: ", df_corrected.iloc[2, 5])

    # DQ Errors
    # 1 - Incompleteness; 2 - Inaccuracy; 3 - Inconsistency; 4 - Duplication

    columns = df.columns.tolist()
    rows = df.fillna("").values.tolist()
    # rows = df.values.tolist()
    # fixed_rows = rows

    incompleteErr = [columns.copy(), [0] * len(columns)]
    inaccurateErr = [columns.copy(), [0] * len(columns)]
    inconsistErr = [columns.copy(), [0] * len(columns)]
    duplicateErr = 0

    incomplete_err_cells = []
    inaccurate_err_cells = []
    inconsist_err_cells = []
    # duplicate_err_rows = []


    # fixed_columns = columns
    # fixed_rows = rows
    # fixed_cells_bool = [[0] * len(columns) for _ in range(len(rows))]
    # print(fixed_cells_bool)

    # Null values
    # resultNull = [[None for _ in range(len(columns))] for _ in range(2)]

    # for i, column in enumerate(columns):
    #     resultNull[0][i] = column
    #     resultNull[1][i] = int(df[column].isnull().sum())

    # print(resultNull)

    # result = [[None for _ in range(len(columns))] for _ in range(2)]


    # Duplication
    duplicates = df[df.duplicated()]
    duplicateErr = len(duplicates)
    duplicate_err_rows = duplicates.index.tolist()

    # for row in duplicate_err_rows:
    #     ws_corrected.delete_rows(row + 2, 1)

    # ID Validation

    # idIndexes = []
    # for column in columns:
    #     if "id" in column.lower():
    #         idIndexes.append(columns.index(column))
    #
    # for idIndex in idIndexes:
    #     uniqueIDs = df.iloc[:, idIndex].nunique()
    #     totalIDs = df.iloc[:, idIndex].count()

    id_indexes = []
    for column in columns:
        cleaned_column = re.sub(r'[0-9]+', '', column.lower())
        if "id" in cleaned_column.split():
            id_indexes.append(columns.index(column))

    for id_index in id_indexes:
        # ids = []
        #
        # for id in df[columns[id_index]].dropna():
        #     ids.append(str(id).strip())

        # ids_series = pd.Series(ids)
        # leading_zero_rate = ids_series.str.match(r'^0+').mean()
        # print("Leading zero rate:", leading_zero_rate)

        # Learn features
        features = []
        for row in rows:
            if row[id_index] is None or len(row[id_index]) == 0:
                incompleteErr[1][id_index] += 1
                incomplete_err_cells.append((rows.index(row) + 2, id_index + 1))
            else:
                cleaned_id = str(row[id_index]).strip()

                length = len(cleaned_id)
                # if len(cleaned_id) > 0:
                #     length = len(cleaned_id)
                # else:
                #     length = len(row[id_index])

                num_digits = len(re.findall(r'\d', cleaned_id))
                digits_ratio = num_digits / length

                num_chars = len(re.findall(r'[A-Za-z]', cleaned_id))
                chars_ratio = num_chars / length
                num_unique_chars = len(set(cleaned_id))

                num_special = len(re.findall(r'[^A-Za-z0-9]', cleaned_id))
                special_ratio = num_special / length

                counts = Counter(cleaned_id)
                probs = [c / len(cleaned_id) for c in counts.values()]
                entropy = -sum(p * math.log2(p) for p in probs)

                # if (entropy > 0.01
                #         and leading_zero_rate < 0.5
                #         and bool(re.match(r'^0+', cleaned_id))):
                #     inconsistErr[1][id_index] += 1
                #     inconsist_err_cells.append((rows.index(row) + 2, id_index + 1))

                # max_repeat_digit = max(counts.values() if type(counts.keys) == int)
                # max_repeat_char = max(counts.values() if type(counts.keys) == str)

                max_repeat_digit = max(
                    (count for p, count in counts.items() if p.isdigit()),
                    default=0
                )

                max_repeat_char = max(
                    (count for p, count in counts.items() if p.isalpha()),
                    default=0
                )

                starts_with_char = int(cleaned_id[0].isalpha())
                ends_with_char = int(cleaned_id[-1].isalpha())

                first_digit_pos = next((i for i, p in enumerate(cleaned_id) if p.isdigit()), length)

                pattern = re.sub(r'[0-9]', '9',re.sub(r'[A-Za-z]', 'A', cleaned_id))
                pattern_code = hash(pattern) % (10 ** 8)

                bigrams = [cleaned_id[i:i + 2] for i in range(len(cleaned_id) - 1)]
                trigrams = [cleaned_id[i:i + 3] for i in range(len(cleaned_id) - 2)]
                n_grams = len(bigrams) + len(trigrams)

                num_unique_bigrams = len(set(bigrams))
                num_unique_trigrams = len(set(trigrams))

                prefix_len = len(re.match(r'^[A-Za-z]+', cleaned_id).group()) \
                    if re.match(r'^[A-Za-z]+', cleaned_id) else 0

                transitions = sum(
                    cleaned_id[i].isalpha() != cleaned_id[i + 1].isalpha()
                    for i in range(len(cleaned_id) - 1)
                )

                num_leading_zeros = len(cleaned_id) - len(cleaned_id.lstrip("0"))
                leading_zero_ratio = num_leading_zeros / len(cleaned_id)


                features.append([
                                 digits_ratio, chars_ratio, leading_zero_ratio,
                                 # max_repeat_digit, max_repeat_char,
                                 # starts_with_char, ends_with_char,
                                 first_digit_pos, prefix_len
                                 ])

        features = np.array(features, dtype=float)
        # print("Features:", features)
        scaler = StandardScaler()
        features_scaled = scaler.fit_transform(features)

        model = IsolationForest(random_state=42)
        model.fit(features_scaled)

        scores = model.decision_function(features_scaled)

        pred = scores < -0.05

        for i, p in enumerate(pred):
            if p == True:
                inconsistErr[1][id_index] += 1
                inconsist_err_cells.append((i + 2, id_index + 1))


    # Name Validation
    name_indexes = []
    for column in columns:
        if "name" in column.lower():
            name_indexes.append(columns.index(column))

    for name_index in name_indexes:

        org_keywords = [
            "Ltd",
            "Limited",
            "Inc",
            "Corporation",
            "Group",
            "House",
            "Home",
            "Care",
            "Centre",
            "University"
        ]

        org_ratio = df[columns[name_index]].dropna().apply(lambda x: any(k.lower() in x.lower() for k in org_keywords)).mean()
        # print(name_index, org_ratio)

        # Person Name
        if any(word in columns[name_index].lower() for word in {"person", "first", "last"}) or org_ratio == 0:
            for row in rows:
                if row[name_index]:
                    valid_name = re.sub(r"[^a-zA-Z\s]", "", row[name_index].title())
                    if row[name_index] != valid_name:
                        inconsistErr[1][name_index] += 1
                        inconsist_err_cells.append((rows.index(row) + 2, name_index + 1))
                        if excel_data:
                            ws_corrected.cell(rows.index(row) + 2, name_index + 1).value = valid_name

        # Organisation Name
        else:
            org_names = df[columns[name_index]].tolist()
            # print(org_names)

            model = SentenceTransformer("all-mpnet-base-v2")
            embeddings = model.encode(org_names)
            similarities = cosine_similarity(embeddings)
            # print("Similarities:", similarities)

            matching_similarities = [org_names.copy(), [None] * len(org_names)]

            for row in rows:
                if row[name_index] is not None:
                    org_name = row[name_index]
                    similar_org_names = []

                    for i, similarity_score in enumerate(similarities[rows.index(row)]):
                        if similarity_score > 0.85 and similarity_score < 0.99:
                            similar_org_names.append(org_names[i])

                    matching_similarities[1][rows.index(row)] = similar_org_names

                    # print("Organisation name:", org_name)
                    # print("Similar organisation names:", similar_org_names)

            # print(matching_similarities)

            for i, sim in enumerate(matching_similarities[1]):
                if sim:
                    inconsistErr[1][name_index] += 1
                    inconsist_err_cells.append((i + 2, name_index + 1))

    # Age Validation
    age_indexes = []
    for column in columns:
        if "age" in column.lower():
            age_indexes.append(columns.index(column))

    for age_index in age_indexes:
        for row in rows:
            age = int(row[age_index])
            if age and (age < 0 or age > 120):
                inaccurateErr[1][age_index] += 1
                inaccurate_err_cells.append((rows.index(row) + 2, age_index + 1))

    # Email Validation
    email_domains = [
        "hillingdon.gov.uk",
        "gmail.com",
        "yahoo.com",
        "hotmail.com",
        "icloud.com"
    ]

    email_indexes = []
    for column in columns:
        if "email" in column.lower():
            email_indexes.append(columns.index(column))

    for email_index in email_indexes:
        for row in rows:
            if row[email_index]:
                email_regex = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
                if not bool(email_regex.fullmatch(row[email_index])):
                    inconsistErr[1][email_index] += 1
                    inconsist_err_cells.append((rows.index(row) + 2, email_index + 1))
                else:
                    email_domain = row[email_index].split("@")[1]
                    for valid_email_domain in email_domains:
                        email_domain_similarity = fuzz.ratio(email_domain.lower(), valid_email_domain.lower())
                        if email_domain_similarity > 85 and email_domain_similarity < 100:
                            inaccurateErr[1][email_index] += 1
                            inaccurate_err_cells.append((rows.index(row) + 2, email_index + 1))
                            if excel_data:
                                ws_corrected.cell(rows.index(row) + 2, email_index + 1).value = (
                                    row[email_index].split("@")[0] + "@" + valid_email_domain)
                            break


    # Address Validation
    LBH_Addresses = pd.read_csv("Full_LBH_Addresses.csv")
    # print("LBH Address", LBH_Addresses.iloc[1, :])

    full_lbh_addresses = LBH_Addresses.iloc[:, 2]

    abbreviation_map = {
        "rd": "road", "road": "road",
        "st": "street", "street": "street",
        "ave": "avenue", "av": "avenue", "avenue": "avenue",
        "ln": "lane", "lane": "lane",
        "dr": "drive", "drive": "drive",
        "pl": "place", "place": "place",
        "ct": "court", "court": "court",
        "gr": "grove", "grove": "grove",
        "cres": "crescent", "crescent": "crescent",
        "cl": "close", "close": "close",
        "terr": "terrace", "terrace": "terrace",
        "gardens": "gardens",
        "way": "way"
    }

    post_towns = [
        "Uxbridge",
        "Hayes",
        "Ruislip",
        "West Drayton",
        "Northwood",
        "Eastcote",
        "Harefield",
        "Pinner",
        "Harlington",
        "Sipson",
        "Ickenham",
        "Heathrow",
        "Northolt"
    ]

    postcode_areas = [
        "UB",
        "HA",
        "TW",
        "WD"
    ]


    UK_POSTCODE_REGEX = re.compile(
        r"\b(GIR 0AA|"
        r"(?:[A-Z]{1,2}[0-9][0-9A-Z]?\s?[0-9][A-Z]{2}))\b",
        re.IGNORECASE
    )

    def extract_postcode(address):
        match = UK_POSTCODE_REGEX.search(address)
        return match.group(0).upper() if match else None

    # Remove extra spaces and apply abbreviation map for consistency
    def normalize_address(address):
        postcode = extract_postcode(address)
        if postcode:
            address = UK_POSTCODE_REGEX.sub("", address).strip(", ")

        parts = address.split(",")
        normalized_address_arr = []

        for part in parts:
            part = part.lower().strip()
            part = re.sub(r"[.]", "", part)

            tokens = part.split()
            normalized_part_arr = []

            for t in tokens:
                # Check if middx is majority
                if t == "middlesex" or t == "middx":
                    normalized_part_arr.append("middlesex")
                else:
                    normalized_part_arr.append(abbreviation_map.get(t, t))

            normalized_address_arr.append(" ".join(normalized_part_arr).title())

        # print("Normalized address: ", normalized_address)

        normalized_address = ", ".join(normalized_address_arr)

        if postcode:
            normalized_address = normalized_address + ", " + postcode

        return normalized_address

    # Reorder based on address hierarchy and manage county
    def reorder_address(address, county_presence_rate):
        postcode = extract_postcode(address)

        if postcode:
            address = UK_POSTCODE_REGEX.sub("", address).strip(", ")

        county = None
        reordered_address_arr = []

        parts = address.split(",")
        for part in parts:
            part = part.lower().strip()

            if part == "middlesex":
                county = part.title()
            else:
                reordered_address_arr.append(part.title())

        if county_presence_rate > 0.5:
            reordered_address_arr.append("Middlesex")

        if postcode:
            reordered_address_arr.append(postcode)

        reordered_address = ", ".join(reordered_address_arr)

        # if (town not in post_towns
        #         and all(post_town not in town for post_town in post_towns)
        #         and postcode
        #         and postcode[:2] in postcode_areas):
        #     print(", ".join(reordered_address_arr))

        return reordered_address

    def check_address(address, row, col):

        initial_address = address

        postcode = extract_postcode(address)
        if postcode:
            address = UK_POSTCODE_REGEX.sub("", address).strip(", ")

        address_arr = []
        parts = address.split(",")
        county = None

        for part in parts:
            part = part.lower().strip()
            if part == "middlesex":
                county = part.title()
            else:
                # if "london" not in part:
                address_arr.append(part.title())

        # address = ", ".join(address_arr)
        # print(address)

        valid_town = None
        # valid_locality = None
        valid_street = None

        # Get valid post town, locality, street by postcode
        if postcode:
            lbh_addresses_same_postcode = []

            for lbh_a in full_lbh_addresses:
                if postcode in lbh_a:
                    lbh_addresses_same_postcode.append(lbh_a)

            # print(lbh_addresses_same_postcode)

            if len(lbh_addresses_same_postcode) > 0:
                all_parts = []

                for lbh_a in lbh_addresses_same_postcode:
                    parts = [p.strip().title() for p in lbh_a.split(",")]
                    parts = parts[:-1]
                    all_parts.append(parts)

                # print(all_parts)

                common_parts = set(all_parts[0])

                for parts in all_parts[1:]:
                    common_parts &= set(parts)

                # print(common_parts)

                sample_parts = all_parts[0]

                sorted_common_parts = sorted(
                    common_parts,
                    key=lambda p: sample_parts.index(p)
                )

                # print(sorted_common_parts)

                valid_town = sorted_common_parts[-1]

                if len(sorted_common_parts) == 2:
                    valid_street = sorted_common_parts[0]

                elif len(sorted_common_parts) == 3:
                    # valid_town = sorted_common_parts[2]

                    # Contains Locality
                    if (not re.search(r'\d',sorted_common_parts[1]) and
                        not re.search(r'\d', sorted_common_parts[0].split()[-1])):
                        valid_street = sorted_common_parts[0]

                    # Contains Building Name
                    else:
                        valid_street = sorted_common_parts[1]

                elif len(sorted_common_parts) == 4:
                    #Contains Building Name and Locality
                    valid_street = sorted_common_parts[1]

        if valid_town:
            valid_town = normalize_address(valid_town)

        if valid_street:
            valid_street = normalize_address(valid_street).strip()
            # valid_street = re.sub(r'\d+', '', valid_street).strip()

        town = address_arr[-1]

        # address = ", ".join(address_arr)
        # if ((valid_town and valid_town not in address and town not in post_towns) or
        #     (valid_street and valid_street not in address)):
        #         print("Invalid address: ", address)
        #         print(valid_town, valid_street)

        if valid_town and valid_town != town and town not in post_towns:
            town_similarity = fuzz.ratio(town.lower(), valid_town.lower())
            # print(town_similarity)
            if town_similarity > 85:
                # print(town, valid_town, "Inaccuracy")
                inaccurateErr[1][col] += 1
                inaccurate_err_cells.append((row + 2, col + 1))

                address_arr.remove(town)
                address_arr.append(valid_town)
                # address = ", ".join(address_arr)
                # print("Fixed address: ", address)
            else:
                # print(town, valid_town, "Incompleteness")
                incompleteErr[1][col] += 1
                incomplete_err_cells.append((row + 2, col + 1))

                address_arr.append(valid_town)
                # address = ", ".join(address_arr)
                # print("Fixed address: ", address)

        if valid_street:
            for i, part in enumerate(address_arr[:-1]):
                part = re.sub(r'\d+', '', part).strip()
                street_similarity = fuzz.ratio(part.lower(), valid_street.lower())

                if street_similarity == 100:
                    break
                else:
                    if street_similarity > 85:
                        # print(part, valid_street, "Inaccuracy")
                        inaccurateErr[1][col] += 1
                        inaccurate_err_cells.append((row + 2, col + 1))

                        address_arr[i] = re.sub(part, valid_street, address_arr[i])
                        break

                    if i == len(address_arr[:-1]) - 1:
                        # print(part, valid_street, "Incompleteness")
                        incompleteErr[1][col] += 1
                        incomplete_err_cells.append((row + 2, col + 1))


        valid_address = ", ".join(
            address_arr +
            ([county] if county else []) +
            ([postcode] if postcode else [])
        )

        # print("Valid address: ", valid_address)

        if initial_address != valid_address:
            if excel_data:
                ws_corrected.cell(row + 2, col + 1).value = valid_address


    address_column_indexes = []
    for column in columns:
        if "address" in column.lower():
            address_column_indexes.append(columns.index(column))

    for address_column_ind in address_column_indexes:
        if is_string_dtype(df[columns[address_column_ind]]):

            addresses = []
            for a in df[columns[address_column_ind]].dropna():
                addresses.append(normalize_address(a))

            addresses_series = pd.Series(addresses)
            county_presence_rate = addresses_series.str.contains(
                "Middlesex",
                case=False,
                na=False
            ).mean()
            # print(county_presence_rate)

            for row in rows:
                if row[address_column_ind]:
                    address = row[address_column_ind]

                    if address != normalize_address(address):
                        # print("Inconsistent(normilisation): ", address)
                        address = normalize_address(address)

                    if address != reorder_address(address, county_presence_rate):
                        # print("Inconsistent(ordering or missing county): ", address)
                        address = reorder_address(address, county_presence_rate)

                    if row[address_column_ind] != address:
                        inconsistErr[1][address_column_ind] += 1
                        inconsist_err_cells.append((rows.index(row) + 2, address_column_ind + 1))
                        if excel_data:
                            ws_corrected.cell(rows.index(row) + 2, address_column_ind + 1).value = address

                    check_address(address, rows.index(row), address_column_ind)


    # Highlight Excel file
    if excel_data:
        wb = load_workbook(excel_data)
        ws = wb.active

        yellow_fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid"
        )

        orange_fill = PatternFill(
            start_color="FF8C00",
            end_color="FF8C00",
            fill_type="solid"
        )

        red_fill = PatternFill(
            start_color="FF4141",
            end_color="FF4141",
            fill_type="solid"
        )

        pink_fill = PatternFill(
            start_color="FF87CF",
            end_color="FF87CF",
            fill_type="solid"
        )

        for row, col in inconsist_err_cells:
            ws.cell(row=row, column=col).fill = yellow_fill

        for row, col in incomplete_err_cells:
            ws.cell(row=row, column=col).fill = orange_fill

        for row, col in inaccurate_err_cells:
            ws.cell(row=row, column=col).fill = red_fill

        for row in duplicate_err_rows:
            for cell in ws[row+2]:
                cell.fill = pink_fill

        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)

        highlighted_excel_sheet = output_stream

    # Get Corrected file

    if fileName.endswith(".xlsx"):
        corrected_stream = io.BytesIO()
        wb_corrected.save(corrected_stream)
        corrected_stream.seek(0)
        corrected_excel_sheet = corrected_stream

    elif fileName.endswith(".csv"):
        corrected_stream = io.BytesIO()
        wb_corrected.save(corrected_stream)
        corrected_stream.seek(0)

        df_corrected = pd.read_excel(corrected_stream, dtype=str)
        csv_stream = io.StringIO()
        df_corrected.to_csv(csv_stream, index=False)
        csv_stream.seek(0)

        corrected_csv_sheet = csv_stream


    return {
        "filename": fileName,
        "columns": columns,
        "rows": rows,
        # "fixed_rows": fixed_rows,
        # "data_types": df.dtypes.astype(str).tolist(),
        # "missing_values": df.isnull().sum().tolist(),
        # "name_errors_count": nameErrorCount,
        # "age_errors_count": ageErrorCount,
        # "result": result,
        # "resultNull": resultNull
        "incompleteErr": incompleteErr,
        "inaccurateErr": inaccurateErr,
        "inconsistErr": inconsistErr,
        "duplicateErr": duplicateErr,
    }

    # csv_reader = csv.reader(csv_data)
    # json_data = list(csv_reader)
    # return {"headers": json_data[0], "data": json_data}

@app.get("/highlighted_file")
async def get_file():
    if highlighted_excel_sheet:
        return StreamingResponse(
            highlighted_excel_sheet,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=highlighted_errors.xlsx",
            }
        )

@app.get("/corrected_file")
async def get_corrected_file():
    if corrected_excel_sheet:
        return StreamingResponse(
            corrected_excel_sheet,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=corrected_errors.xlsx",
            }
        )
    elif corrected_csv_sheet:
        return StreamingResponse(
            corrected_csv_sheet,
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=corrected_errors.csv",
            }
        )


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
